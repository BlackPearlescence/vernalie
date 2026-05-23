from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


OUTPUT_DIR = Path(__file__).resolve().parents[2] / "seed-data" / "import-fixtures"

WORKBOOKS = [
    {
        "file_name": "northern-rootstock-mixed.xlsx",
        "sheet_name": "Spring Inventory",
        "title": "Northern Rootstock Cooperative - Spring Inventory",
        "headers": [
            "Species",
            "Named Variety",
            "Units Started",
            "Location Code",
            "Structure",
            "Planting Date",
            "Batch Status",
            "Shock Factor",
            "Contract Ref",
            "Committed",
            "Ship Date",
        ],
        "rows": [
            ["Apple", "Liberty", 900, "NR-A1", "OPEN_FIELD", date(2026, 1, 12), "IN_PRODUCTION", 0.01, "NRC-400", 760, date(2027, 3, 25)],
            ["Apple", "Enterprise", 820, "NR-A2", "OPEN_FIELD", date(2026, 1, 14), "IN_PRODUCTION", 0.02, "NRC-401", 700, date(2027, 3, 25)],
            ["Cherry", "Montmorency", 650, "NR-C1", "OPEN_FIELD", date(2026, 2, 1), "IN_PRODUCTION", 0.03, "NRC-402", 560, date(2027, 4, 20)],
            ["Blueberry", "Bluecrop", 1200, "NR-B4", "SHADE_HOUSE", date(2026, 2, 15), "IN_PRODUCTION", 0, "NRC-403", 950, date(2027, 5, 5)],
            ["Strawberry", "Albion", 2500, "Runner Bay 1", "GREENHOUSE", date(2026, 3, 1), "IN_PRODUCTION", 0, "NRC-404", 2200, date(2026, 6, 15)],
            ["Raspberry", "Heritage", 1400, "Cane Row 3", "OPEN_FIELD", date(2026, 3, 5), "ON_HOLD", 0.05, None, 0, date(2027, 5, 20)],
        ],
    },
    {
        "file_name": "suncrest-annuals-plugs.xlsx",
        "sheet_name": "Plug Runs",
        "title": "Suncrest Plug Farm - Annuals and Bedding Plants",
        "headers": [
            "Item",
            "Cultivar Name",
            "Tray Count",
            "Growing Area",
            "House Type",
            "Seeded",
            "Stage",
            "Loss Override",
            "PO",
            "Promise Qty",
            "Ready Week",
        ],
        "rows": [
            ["Marigold", "Inca II Orange", 6400, "House 1 Bench A", "HIGH_TECH_GREENHOUSE", date(2026, 3, 4), "IN_PRODUCTION", 0, "PO-771", 5800, date(2026, 5, 1)],
            ["Petunia", "Supertunia Vista Bubblegum", 5200, "House 1 Bench B", "HIGH_TECH_GREENHOUSE", date(2026, 3, 6), "IN_PRODUCTION", 0.01, "PO-772", 4600, date(2026, 5, 10)],
            ["Zinnia", "Benary Giant Mix", 4800, "House 2 Bench C", "GREENHOUSE", date(2026, 3, 10), "IN_PRODUCTION", 0, "PO-773", 4200, date(2026, 5, 15)],
            ["Pansy", "Matrix Blue Blotch", 3600, "House 2 Bench D", "GREENHOUSE", date(2026, 2, 15), "READY_TO_SHIP", 0, "PO-774", 3200, date(2026, 4, 10)],
            ["Salvia", "Victoria Blue", 3100, "House 3 Bench A", "GREENHOUSE", date(2026, 3, 12), "IN_PRODUCTION", 0.02, "PO-775", 2600, date(2026, 5, 20)],
            ["Geranium", "Maverick Red", 2800, "House 3 Bench B", "GREENHOUSE", date(2026, 2, 25), "IN_PRODUCTION", 0, None, 0, date(2026, 5, 5)],
        ],
    },
]


def build_workbook(spec):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = spec["sheet_name"]
    sheet.append([spec["title"]])
    sheet.append(spec["headers"])

    for row in spec["rows"]:
        sheet.append(row)

    sheet["A1"].font = Font(bold=True, size=14)
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF1E8")

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 28)

    for row in sheet.iter_rows(min_row=3, min_col=6, max_col=6):
        row[0].number_format = "yyyy-mm-dd"

    for row in sheet.iter_rows(min_row=3, min_col=11, max_col=11):
        row[0].number_format = "yyyy-mm-dd"

    workbook.save(OUTPUT_DIR / spec["file_name"])


OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
for workbook_spec in WORKBOOKS:
    build_workbook(workbook_spec)
